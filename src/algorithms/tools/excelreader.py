from openpyxl import Workbook, load_workbook
from entities.group import Group
from entities.user import User

#Used for testing real data. Only works for the single excel file at the moment. (Kohti tutkivaa työtapaa VO -22 SiltaSanoma lööpit (jakoon))

def create_groups(excel):
    '''Takes in the name of the excel chart, returns the list of all possible groups.'''
    book = load_workbook(excel)
    sheet = book.active
    groups = {}
    list = sheet['Q2'].value.split(';')
    del list[-1]
    for i, l in enumerate(list):
        if l != '':
            groups[i] = Group(i, l, 12)
    return groups

def create_users(excel, groups):
    '''Takes in the name of the excel chart and the list of groups. Returns the list of users with the list of their selections'''
    book = load_workbook(excel)
    sheet = book.active
    users = {}

    listofselections = sheet['Q']
    i = 0
    for l in listofselections:
        if l.value == 'Lööpit kiinnostusjärjestykseen!': continue
        selections = l.value.split(';')
        selection_list = []
        for s in selections:
            for g in groups:
                if groups[g].name == s:
                    selection_list.append(groups[g].id)
        users[i] = User(i, 'opiskelija' + str(i), selection_list)
        i += 1
    return users
