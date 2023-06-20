from openpyxl import load_workbook
from entities.group import Group
from entities.student import Student

#Used for testing real data. Only works for the single excel file at the moment. (Kohti tutkivaa työtapaa VO -22 SiltaSanoma lööpit (jakoon))
EXCEL = "tutki.xlsx"

def create_groups():
    '''Takes in the name of the excel chart, returns the list of all possible groups.'''
    book = load_workbook(EXCEL)
    sheet = book.active
    groups = {}
    list = sheet['Q2'].value.split(';')
    for i, list_value in enumerate(list):
        if list_value != '':
            groups[i] = Group(i, list_value, 11)
    return groups

def create_students(groups):
    '''Takes in the name of the excel chart and the list of groups. Returns the list of users with the list of their selections'''
    book = load_workbook(EXCEL)
    sheet = book.active
    students = {}

    listofselections = sheet['Q']
    i = 0
    for list_value in listofselections: # pylint: disable=R1702
        if list_value.value == 'Lööpit kiinnostusjärjestykseen!': continue
        selections = list_value.value.split(';')
        selection_list = []
        for selection in selections:
            for group_id in groups:
                if groups[group_id].name == selection:
                    selection_list.append(groups[group_id].id)
        students[i] = Student(i, 'opiskelija' + str(i), selection_list)
        i += 1
    return students
